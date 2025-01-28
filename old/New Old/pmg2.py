#!/usr/bin/env python3
"""
pmg.py

Extension of your reverted code that:
1) Reads DMS (pinnacle columns) + website sets/dicts.
2) Builds "master" DataFrame (DMS + website-only),
   splits into separate dealer sheets (in DMS), a `to_be_removed` sheet (website-only),
   and now a new `to_dos` sheet with auto-generated tasks.
3) Adds color-coded conditional formatting on each prefix sheet
   and on `to_be_removed` (green if all 3 sites = "Yes", red if PhotoCount>1 & any site="No").
4) The `to_dos` sheet is also a table with auto-sized columns.

Usage:
  python pmg.py
"""

import os
import warnings
import pandas as pd

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utilities import (
    read_csv_with_sep_check,
    clean_dataframe,
    col_index_to_excel_col_name
)

# Suppress openpyxl's "no default style" warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ------------------- CONFIG -------------------
PINNACLE_SCHEMA = [
    "Stock Number",
    "Make",
    "Model",
    "Specification",
    "Colour",
    "Registration Date",
    "VIN",
    "Odometer",
    "Photo Count",
    "Selling Price",
    "Stand In Value",
    "Internet Price",
    "Date In Stock",
    "Original Group Date In Stock",
    "Stock Days",
    "Branch",
    "Location",
    "Body Style",
    "Fuel Type",
    "Transmission",
    "Customer Ordered",
    "Profiles",
]

DEALER_PREFIXES = {
    "Ford_Nelspruit":   "UF",
    "Ford_Mazda":       "UG",
    "Produkta_Nissan":  "UA",
    "Suzuki_Nelspruit": "UE",  # forced is_on_cars="Yes"
    "Ford_Malalane":    "US",
}

# ------------------ STEP A: READ DMS INTO DICT ------------------
def extract_dms_dict(dms_file: str) -> dict:
    """Reads pmg_dms_data.csv -> {stockNum -> rowData} with only PINNACLE_SCHEMA."""
    if not os.path.isfile(dms_file):
        print(f"[WARN] Missing DMS file: {dms_file}")
        return {}

    temp_cols = pd.read_csv(dms_file, nrows=1).columns.tolist()
    wanted_cols = []
    for col in PINNACLE_SCHEMA:
        if col == "Customer Ordered" and "Customer Order" in temp_cols:
            wanted_cols.append("Customer Order")
        elif col in temp_cols:
            wanted_cols.append(col)

    df = pd.read_csv(dms_file, usecols=wanted_cols)
    if "Customer Order" in df.columns:
        df.rename(columns={"Customer Order": "Customer Ordered"}, inplace=True)

    # Reorder & strip
    final_cols = [c for c in PINNACLE_SCHEMA if c in df.columns]
    df = df[final_cols]
    for c in df.select_dtypes(include=['object']).columns:
        df[c] = df[c].astype(str).str.strip()

    # Convert to dict
    dms_map = {}
    if "Stock Number" not in df.columns:
        return dms_map

    for _, row in df.iterrows():
        sn = row["Stock Number"]
        if sn:
            dms_map[sn] = dict(row)
    return dms_map

# ------------------ STEP B: READ WEBSITE DATA ------------------
def read_autotrader_data(src_folder: str):
    """
    Reads autotrader.csv from each dealership subfolder.
    """
    stks, prices = set(), {}
    for sub in os.listdir(src_folder):
        path_sub = os.path.join(src_folder, sub)
        if not os.path.isdir(path_sub):
            continue
        csv_file = os.path.join(path_sub, "autotrader.csv")
        if not os.path.isfile(csv_file):
            continue

        df = read_csv_with_sep_check(csv_file)
        df.columns = df.columns.str.strip()
        if "StockNumber" in df.columns:
            df.rename(columns={"StockNumber": "Stock Number"}, inplace=True)

        price_col = next((c for c in ["PriceFormatted", "Price", "price"] if c in df.columns), None)

        if "Stock Number" in df.columns:
            for _, row in df.iterrows():
                sn = str(row["Stock Number"]).strip()
                if sn:
                    stks.add(sn)
                    p = str(row[price_col]).strip() if price_col else ""
                    prices[sn] = p
    return stks, prices


def read_cars_data(src_folder: str):
    """
    Reads cars.xlsx from each dealership subfolder.
    """
    stks, prices = set(), {}
    for sub in os.listdir(src_folder):
        path_sub = os.path.join(src_folder, sub)
        if not os.path.isdir(path_sub):
            continue
        xlsx_file = os.path.join(path_sub, "cars.xlsx")
        if not os.path.isfile(xlsx_file):
            continue

        df = pd.read_excel(xlsx_file)
        df.columns = df.columns.str.strip()
        if "Reference" in df.columns:
            df.rename(columns={"Reference": "Stock Number"}, inplace=True)

        price_col = next((c for c in ["Price", "price"] if c in df.columns), None)

        if "Stock Number" in df.columns:
            for _, row in df.iterrows():
                sn = str(row["Stock Number"]).strip()
                if sn:
                    stks.add(sn)
                    p = str(row[price_col]).strip() if price_col else ""
                    prices[sn] = p
    return stks, prices


def read_pmg_web_data(src_folder: str):
    """
    Reads pmg_web_data.csv if present in src/.
    """
    csv_file = os.path.join(src_folder, "pmg_web_data.csv")
    if not os.path.isfile(csv_file):
        return set(), {}

    df = read_csv_with_sep_check(csv_file)
    df.columns = df.columns.str.strip()
    if "SKU" in df.columns:
        df.rename(columns={"SKU": "Stock Number"}, inplace=True)

    price_col = next((c for c in ["Regular price", "Regular Price", "price", "Price"] if c in df.columns), None)

    stks, prices = set(), {}
    if "Stock Number" in df.columns:
        for _, row in df.iterrows():
            sn = str(row["Stock Number"]).strip()
            if sn:
                stks.add(sn)
                p = str(row[price_col]).strip() if price_col else ""
                prices[sn] = p
    return stks, prices

# ------------------ STEP C: BUILD MASTER DATAFRAME ------------------
def build_master_dataframe(dms_map: dict,
                           at_set: set, at_prices: dict,
                           cars_set: set, cars_prices: dict,
                           pmg_set: set, pmg_prices: dict) -> pd.DataFrame:
    """
    Union of DMS + website sets. Each row => in_dms, site columns, etc.
    If prefix=UE => is_on_cars="Yes".
    """
    all_stocks = set(dms_map.keys()) | at_set | cars_set | pmg_set
    rows = []

    for sn in sorted(all_stocks):
        row = {}
        # in DMS or not
        if sn in dms_map:
            row.update(dms_map[sn])
            row["in_dms"] = True
        else:
            for c in PINNACLE_SCHEMA:
                row[c] = ""
            row["in_dms"] = False

        # Cars
        if sn in cars_set:
            row["is_on_cars"] = "Yes"
            row["cars_price"] = cars_prices.get(sn, "")
        else:
            row["is_on_cars"] = "No"
            row["cars_price"] = ""

        # AutoTrader
        if sn in at_set:
            row["is_on_autotrader"] = "Yes"
            row["autotrader_price"] = at_prices.get(sn, "")
        else:
            row["is_on_autotrader"] = "No"
            row["autotrader_price"] = ""

        # pmg web
        if sn in pmg_set:
            row["is_on_pmgWeb"] = "Yes"
            row["pmg_web_price"] = pmg_prices.get(sn, "")
        else:
            row["is_on_pmgWeb"] = "No"
            row["pmg_web_price"] = ""

        # Force is_on_cars="Yes" if prefix=UE
        if sn.startswith("UE"):
            row["is_on_cars"] = "Yes"

        row["Stock Number"] = sn
        rows.append(row)

    return pd.DataFrame(rows)


def reorder_final_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Put main Pinnacle columns up front (plus 'in_dms'),
    then Photo Count, Internet Price, site columns at the end.
    """
    front_cols = [
        "Stock Number", "Make", "Model", "Specification", "Colour",
        "Registration Date", "VIN", "Odometer", "Selling Price",
        "Stand In Value", "Date In Stock", "Original Group Date In Stock",
        "Stock Days", "Branch", "Location", "Body Style", "Fuel Type",
        "Transmission", "Customer Ordered", "Profiles", "in_dms"
    ]
    end_cols = [
        "Photo Count", "Internet Price",
        "is_on_cars", "cars_price",
        "is_on_autotrader", "autotrader_price",
        "is_on_pmgWeb", "pmg_web_price"
    ]

    existing = df.columns.tolist()
    final_order = [c for c in front_cols if c in existing] + [c for c in end_cols if c in existing]
    df = df.reindex(columns=final_order)
    return clean_dataframe(df)

# -------------- NEW: Generate to_dos --------------
def generate_todos(df_in_dms: pd.DataFrame, df_removed: pd.DataFrame) -> pd.DataFrame:
    """
    Auto-generate tasks:
     - If PhotoCount>1 & any site=="No" => "Need to fix listing"
     - If website-only => "Remove from site"
    """
    tasks = []

    # Condition1: In DMS, Photo>1 & any site col == "No"
    if not df_in_dms.empty and "Photo Count" in df_in_dms.columns:
        for _, row in df_in_dms.iterrows():
            pc = row.get("Photo Count", "0")
            try:
                pc_val = float(pc)
            except:
                pc_val = 0
            if pc_val>1:
                # check site columns
                for sitecol in ["is_on_cars","is_on_autotrader","is_on_pmgWeb"]:
                    if row.get(sitecol,"No")=="No":
                        tasks.append({
                            "Task":"Need to fix listing",
                            "Stock Number": row["Stock Number"],
                            "Notes":f"PhotoCount>1 but missing on {sitecol}"
                        })
                        break

    # Condition2: website-only => remove from site
    if not df_removed.empty:
        for _, row in df_removed.iterrows():
            tasks.append({
                "Task":"Remove from site",
                "Stock Number": row["Stock Number"],
                "Notes":"Website-only"
            })

    df_tasks = pd.DataFrame(tasks, columns=["Task","Stock Number","Notes"])
    return df_tasks


# -------------- STEP D: STYLING HELPERS --------------
def auto_size_columns(sheet, df: pd.DataFrame):
    """
    Approximate "auto size" for each column using sample.
    """
    from openpyxl.utils import get_column_letter

    for col_idx, col_name in enumerate(df.columns, start=1):
        header_len = len(col_name)
        sample_values = df[col_name].astype(str).head(50).tolist()
        avg_len = sum(len(v) for v in sample_values)/max(len(sample_values),1)
        best_len = int(max(header_len, avg_len)) + 2
        sheet.column_dimensions[get_column_letter(col_idx)].width = best_len


def create_excel_table(sheet, df: pd.DataFrame, table_name="DataTable"):
    """
    Insert an Excel Table with TableStyleMedium9.
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    rows = df.shape[0]
    cols = df.shape[1]
    if rows<1 or cols<1:
        return

    last_col_letter = col_index_to_excel_col_name(cols - 1)
    ref = f"A1:{last_col_letter}{rows+1}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)


def apply_conditional_formatting(sheet, df: pd.DataFrame):
    """
    1) Green fill if all site columns == "Yes"
    2) Red fill if Photo Count>1 & any site col== "No"
    """
    all_cols = df.columns.tolist()

    # 1) Green fill => all site == Yes
    try:
        cars_ix = all_cols.index("is_on_cars")
        auto_ix = all_cols.index("is_on_autotrader")
        pmg_ix  = all_cols.index("is_on_pmgWeb")
    except ValueError:
        pass
    else:
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        c_cars = col_index_to_excel_col_name(cars_ix)
        c_auto = col_index_to_excel_col_name(auto_ix)
        c_pmg  = col_index_to_excel_col_name(pmg_ix)
        max_row = len(df) + 1
        max_col_letter = col_index_to_excel_col_name(len(all_cols) - 1)
        rng = f"A2:{max_col_letter}{max_row}"
        formula_green = f'AND(${c_cars}2="Yes", ${c_auto}2="Yes", ${c_pmg}2="Yes")'
        rule_green = FormulaRule(formula=[formula_green], fill=green_fill, stopIfTrue=False)
        sheet.conditional_formatting.add(rng, rule_green)

    # 2) Red fill => Photo Count>1 & any site=="No"
    if "Photo Count" in all_cols and "is_on_cars" in all_cols and "is_on_autotrader" in all_cols and "is_on_pmgWeb" in all_cols:
        photo_ix = all_cols.index("Photo Count")
        c_photo = col_index_to_excel_col_name(photo_ix)
        c_cars = col_index_to_excel_col_name(all_cols.index("is_on_cars"))
        c_auto = col_index_to_excel_col_name(all_cols.index("is_on_autotrader"))
        c_pmg  = col_index_to_excel_col_name(all_cols.index("is_on_pmgWeb"))
        max_row = len(df) + 1
        max_col_letter = col_index_to_excel_col_name(len(all_cols) - 1)
        rng = f"A2:{max_col_letter}{max_row}"

        formula_red = f'AND(${c_photo}2>1, OR(${c_cars}2="No", ${c_auto}2="No", ${c_pmg}2="No"))'
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        rule_red = FormulaRule(formula=[formula_red], fill=red_fill, stopIfTrue=False)
        sheet.conditional_formatting.add(rng, rule_red)


def style_sheet(writer, sheet_name, df: pd.DataFrame):
    """
    Create table, auto-size columns, apply color rules.
    """
    sheet = writer.sheets[sheet_name]
    create_excel_table(sheet, df, table_name=f"{sheet_name}Table")
    auto_size_columns(sheet, df)
    apply_conditional_formatting(sheet, df)

# -------------- STEP E: MAIN --------------
def main():
    src_folder = "src"
    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    # 1) DMS => dict
    dms_file = os.path.join(src_folder, "pmg_dms_data.csv")
    dms_map = extract_dms_dict(dms_file)

    # 2) Website sets/dicts
    at_set, at_prices   = read_autotrader_data(src_folder)
    cars_set, cars_prices = read_cars_data(src_folder)
    pmg_set, pmg_prices   = read_pmg_web_data(src_folder)

    # 3) Build master
    df_master = build_master_dataframe(
        dms_map, at_set, at_prices, cars_set, cars_prices, pmg_set, pmg_prices
    )
    df_master = reorder_final_columns(df_master)

    # 4) Split => in_dms vs. to_be_removed
    recognized_prefixes = tuple(DEALER_PREFIXES.values())

    df_to_remove = df_master[
        (df_master["in_dms"]==False) &
        df_master["Stock Number"].str.startswith(recognized_prefixes, na=False)
    ].copy()

    df_in_dms = df_master[
        (df_master["in_dms"]==True) &
        df_master["Stock Number"].str.startswith(recognized_prefixes, na=False)
    ].copy()

    # Create a to_dos from the logic
    df_todos = generate_todos(df_in_dms, df_to_remove)

    # "to_be_removed" => 6 columns only
    remove_cols = [
        "Stock Number",
        "is_on_cars", "cars_price",
        "is_on_autotrader", "autotrader_price",
        "is_on_pmgWeb"
    ]
    df_to_remove = df_to_remove.reindex(columns=remove_cols)
    df_to_remove = clean_dataframe(df_to_remove)

    # 5) Write sheets
    out_file = os.path.join(output_folder, "dealers_by_prefix.xlsx")
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        sheet_count = 0

        # A) One sheet per recognized prefix (from df_in_dms)
        for dealer_name, prefix in DEALER_PREFIXES.items():
            df_sub = df_in_dms[df_in_dms["Stock Number"].str.startswith(prefix)].copy()
            if df_sub.empty:
                continue
            df_sub.to_excel(writer, sheet_name=dealer_name, index=False)
            sheet_count += 1

        # B) Write 'to_be_removed'
        if not df_to_remove.empty:
            df_to_remove.to_excel(writer, sheet_name="to_be_removed", index=False)
            sheet_count += 1

        # C) Write 'to_dos'
        if not df_todos.empty:
            df_todos.to_excel(writer, sheet_name="to_dos", index=False)
            sheet_count += 1

        # Style each sheet
        for sheet_name in writer.sheets.keys():
            if sheet_name == "to_be_removed":
                style_sheet(writer, sheet_name, df_to_remove)
            elif sheet_name == "to_dos":
                # no site columns => skip color coding, but do table + auto-size
                sheet = writer.sheets[sheet_name]
                create_excel_table(sheet, df_todos, table_name="to_dosTable")
                auto_size_columns(sheet, df_todos)
            else:
                # must be a dealer prefix
                prefix = DEALER_PREFIXES.get(sheet_name, None)
                if prefix:
                    df_sub = df_in_dms[df_in_dms["Stock Number"].str.startswith(prefix)].copy()
                    style_sheet(writer, sheet_name, df_sub)

    print(f"[INFO] Wrote {sheet_count} sheets to {out_file} with styling, table, color-coded columns.")
    print("One sheet per dealer prefix, plus 'to_be_removed' & 'to_dos'.")

if __name__ == "__main__":
    main()
