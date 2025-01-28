#!/usr/bin/env python3
"""
stockgpt.py

Fixes:
 - Creates one sheet per dealership prefix for in-DMS vehicles
 - Ensures green highlight if all sites == "Yes"
 - Ensures red highlight if PhotoCount>1 & any site == "No"
 - Reads autotrader.csv in each subfolder (and cars.xlsx, pmg_web_data)
 - 'to_be_removed' & 'to_dos' are separate sheets, each in Excel table form
 - Auto-sizes columns in each sheet
"""

import os
import pandas as pd
import warnings

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------- CONFIG ----------------------
SRC_FOLDER = "src"
OUTPUT_FOLDER = "output"
OUTPUT_FILE = "stockgpt.xlsx"

# Recognized prefixes & their "dealer" names
DEALER_PREFIXES = {
    "Ford_Nelspruit":   "UF",  # e.g. 'UFxxxx'
    "Ford_Mazda":       "UG",
    "Produkta_Nissan":  "UA",  # Nissan
    "Suzuki_Nelspruit": "UE",  # Suzuki => forced is_on_cars="Yes"
    "Ford_Malalane":    "US",
}

PINNACLE_SCHEMA = [
    "Stock Number",
    "Make",
    "Model",
    "Specification",
    "Colour",
    "Registration Date",
    "VIN",
    "Odometer",
    "Photo Count",
    "Selling Price",
    "Stand In Value",
    "Internet Price",
    "Date In Stock",
    "Original Group Date In Stock",
    "Stock Days",
    "Branch",
    "Location",
    "Body Style",
    "Fuel Type",
    "Transmission",
    "Customer Ordered",
    "Profiles",
]


# ---------------------- DATA READING ----------------------
def read_dms_data() -> pd.DataFrame:
    """
    Reads pmg_dms_data.csv from src/, extracting only PINNACLE_SCHEMA columns.
    Renames 'Customer Order' -> 'Customer Ordered' if needed.
    """
    path = os.path.join(SRC_FOLDER, "pmg_dms_data.csv")
    if not os.path.isfile(path):
        print(f"[WARN] {path} not found. Returning empty DMS DataFrame.")
        return pd.DataFrame(columns=PINNACLE_SCHEMA)

    df_temp = pd.read_csv(path, nrows=1)
    actual_cols = df_temp.columns.tolist()
    wanted = []
    for col in PINNACLE_SCHEMA:
        if col == "Customer Ordered" and "Customer Order" in actual_cols:
            wanted.append("Customer Order")
        elif col in actual_cols:
            wanted.append(col)

    if not wanted:
        print("[WARN] None of the Pinnacle columns found in pmg_dms_data.csv.")
        return pd.DataFrame(columns=PINNACLE_SCHEMA)

    df = pd.read_csv(path, usecols=wanted)
    if "Customer Order" in df.columns:
        df.rename(columns={"Customer Order": "Customer Ordered"}, inplace=True)

    # Reorder
    final_cols = [c for c in PINNACLE_SCHEMA if c in df.columns]
    df = df[final_cols]
    # Trim whitespace
    obj_cols = df.select_dtypes(include=['object']).columns
    for c in obj_cols:
        df[c] = df[c].astype(str).str.strip()
    return df


def read_autotrader_data() -> (set, dict):
    """
    Scans subfolders in src/ for autotrader.csv,
    building a set_of_stockNumbers & dict_of_price => {stockNumber: priceStr}
    If 'StockNumber' => rename -> 'Stock Number'
    If 'PriceFormatted'/'Price' => store in price dict
    """
    stks = set()
    prices = {}
    for folder_name in os.listdir(SRC_FOLDER):
        folder_path = os.path.join(SRC_FOLDER, folder_name)
        if not os.path.isdir(folder_path):
            continue

        csv_path = os.path.join(folder_path, "autotrader.csv")
        if not os.path.isfile(csv_path):
            continue

        df = pd.read_csv(csv_path)
        df.columns = df.columns.str.strip()
        if "StockNumber" in df.columns:
            df.rename(columns={"StockNumber": "Stock Number"}, inplace=True)

        price_col = None
        for c in ["PriceFormatted", "Price", "price"]:
            if c in df.columns:
                price_col = c
                break

        if "Stock Number" in df.columns:
            for _, row in df.iterrows():
                sn = str(row["Stock Number"]).strip()
                if sn:
                    stks.add(sn)
                    p_str = str(row[price_col]).strip() if price_col else ""
                    prices[sn] = p_str
    return stks, prices


def read_cars_data() -> (set, dict):
    """
    Scans subfolders for cars.xlsx,
    returning (set_of_stockNums, dict_of_prices)
    If 'Reference' => rename -> 'Stock Number'
    If 'Price' => store in dict
    """
    stks = set()
    prices = {}
    for folder_name in os.listdir(SRC_FOLDER):
        folder_path = os.path.join(SRC_FOLDER, folder_name)
        if not os.path.isdir(folder_path):
            continue

        xlsx_path = os.path.join(folder_path, "cars.xlsx")
        if not os.path.isfile(xlsx_path):
            continue

        df = pd.read_excel(xlsx_path)
        df.columns = df.columns.str.strip()

        if "Reference" in df.columns:
            df.rename(columns={"Reference": "Stock Number"}, inplace=True)

        price_col = None
        for c in ["Price", "price"]:
            if c in df.columns:
                price_col = c
                break

        if "Stock Number" in df.columns:
            for _, row in df.iterrows():
                sn = str(row["Stock Number"]).strip()
                if sn:
                    stks.add(sn)
                    prices[sn] = str(row[price_col]).strip() if price_col else ""
    return stks, prices


def read_pmg_web_data() -> (set, dict):
    """
    Reads pmg_web_data.csv if present => (set_of_sn, dict_of_prices).
    If 'SKU' => rename -> 'Stock Number'.
    If 'Regular price' etc => store in dict.
    """
    path = os.path.join(SRC_FOLDER, "pmg_web_data.csv")
    if not os.path.isfile(path):
        return set(), {}

    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()
    if "SKU" in df.columns:
        df.rename(columns={"SKU": "Stock Number"}, inplace=True)

    price_col = None
    for c in ["Regular price", "Regular Price", "price", "Price"]:
        if c in df.columns:
            price_col = c
            break

    stks = set()
    prices = {}
    if "Stock Number" in df.columns:
        for _, row in df.iterrows():
            sn = str(row["Stock Number"]).strip()
            if sn:
                stks.add(sn)
                prices[sn] = str(row[price_col]).strip() if price_col else ""
    return stks, prices


# ---------------------- MASTER LOGIC ----------------------
def build_master_df(df_dms: pd.DataFrame,
                    at_stk, at_prices,
                    cars_stk, cars_prices,
                    pmg_stk, pmg_prices) -> pd.DataFrame:
    """
    Union of DMS + website sets.
    If prefix=UE => is_on_cars="Yes" forced
    """
    # Convert DMS to dict {sn => row}
    dms_dict = {}
    if "Stock Number" in df_dms.columns:
        for _, row in df_dms.iterrows():
            sn = str(row["Stock Number"]).strip()
            if sn:
                dms_dict[sn] = row.to_dict()

    all_sn = set(dms_dict.keys()) | at_stk | cars_stk | pmg_stk
    rows = []
    for sn in sorted(all_sn):
        row_data = {}
        # in_dms or not
        if sn in dms_dict:
            row_data.update(dms_dict[sn])
            row_data["in_dms"] = True
        else:
            # fill pinnacle with blank
            for c in PINNACLE_SCHEMA:
                row_data[c] = ""
            row_data["in_dms"] = False

        # site columns
        row_data["is_on_cars"] = "Yes" if sn in cars_stk else "No"
        row_data["cars_price"] = cars_prices.get(sn, "")

        row_data["is_on_autotrader"] = "Yes" if sn in at_stk else "No"
        row_data["autotrader_price"] = at_prices.get(sn, "")

        row_data["is_on_pmgWeb"] = "Yes" if sn in pmg_stk else "No"
        row_data["pmg_web_price"] = pmg_prices.get(sn, "")

        # Force is_on_cars if prefix=UE
        if sn.startswith("UE"):
            row_data["is_on_cars"] = "Yes"

        row_data["Stock Number"] = sn
        rows.append(row_data)

    df = pd.DataFrame(rows)
    # reorder columns
    # define a reorder function
    front_cols = [
        "Stock Number", "Make", "Model", "Specification", "Colour",
        "Registration Date", "VIN", "Odometer", "Selling Price",
        "Stand In Value", "Date In Stock", "Original Group Date In Stock",
        "Stock Days", "Branch", "Location", "Body Style", "Fuel Type",
        "Transmission", "Customer Ordered", "Profiles", "in_dms"
    ]
    end_cols = [
        "Photo Count", "Internet Price",
        "is_on_cars", "cars_price",
        "is_on_autotrader", "autotrader_price",
        "is_on_pmgWeb", "pmg_web_price"
    ]
    existing = df.columns.tolist()
    final_order = [c for c in front_cols if c in existing] + [c for c in end_cols if c in existing]
    df = df.reindex(columns=final_order)

    # fill strings
    obj_cols = df.select_dtypes(include=['object']).columns
    for c in obj_cols:
        df[c] = df[c].fillna("")

    return df


def color_code_worksheet(ws, df):
    """
    Color-code:
      - Green if all 3 site cols == "Yes"
      - Red if PhotoCount>1 & any site == "No"
    """
    # Range from A2 => lastcol => lastrow
    rows = len(df)
    cols = len(df.columns)
    if rows<1 or cols<1:
        return

    from openpyxl.utils import get_column_letter

    last_col_letter = get_column_letter(cols)
    rng = f"A2:{last_col_letter}{rows+1}"  # +1 for header

    # 1) green rule
    try:
        c_cars = df.columns.get_loc("is_on_cars")+1   # 1-based
        c_auto = df.columns.get_loc("is_on_autotrader")+1
        c_pmg  = df.columns.get_loc("is_on_pmgWeb")+1
    except KeyError:
        pass
    else:
        cC = get_column_letter(c_cars)
        cA = get_column_letter(c_auto)
        cP = get_column_letter(c_pmg)
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        formula_green = f'AND(${cC}2="Yes", ${cA}2="Yes", ${cP}2="Yes")'
        ruleG = FormulaRule(formula=[formula_green], fill=fill_green, stopIfTrue=False)
        ws.conditional_formatting.add(rng, ruleG)

    # 2) red rule => Photo Count>1 & (any site==No)
    if "Photo Count" in df.columns and "is_on_cars" in df.columns and \
       "is_on_autotrader" in df.columns and "is_on_pmgWeb" in df.columns:
        pc_ix = df.columns.get_loc("Photo Count")+1
        c_cars = df.columns.get_loc("is_on_cars")+1
        c_auto = df.columns.get_loc("is_on_autotrader")+1
        c_pmg  = df.columns.get_loc("is_on_pmgWeb")+1

        pcC = get_column_letter(pc_ix)
        cc  = get_column_letter(c_cars)
        ca  = get_column_letter(c_auto)
        cp  = get_column_letter(c_pmg)

        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        formula_red = f'AND(${pcC}2>1, OR(${cc}2="No", ${ca}2="No", ${cp}2="No"))'
        ruleR = FormulaRule(formula=[formula_red], fill=fill_red, stopIfTrue=False)
        ws.conditional_formatting.add(rng, ruleR)


def auto_size_and_table(ws, df, table_name="DataTable"):
    """
    Auto-size columns based on header+sample data, then
    convert the data region to an Excel Table (for sorting).
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    rows = df.shape[0]
    cols = df.shape[1]
    if rows<1 or cols<1:
        return

    # Table reference => from A1 to last column/row
    last_col_letter = get_column_letter(cols)
    ref = f"A1:{last_col_letter}{rows+1}"  # +1 for header row

    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # auto-size columns
    # sample up to 50 rows for each column
    for col_idx, col_name in enumerate(df.columns, start=1):
        header_len = len(col_name)
        sample_vals = df[col_name].astype(str).head(50).tolist()
        avg_len = sum(len(x) for x in sample_vals)/max(len(sample_vals),1)
        best_len = int(max(header_len, avg_len))+2
        ws.column_dimensions[get_column_letter(col_idx)].width = best_len


def generate_todos(df_in, df_rm) -> pd.DataFrame:
    """
    Create tasks:
      1) if PhotoCount>1 & any site=No => 'Need to fix listing'
      2) if row in df_rm => 'Remove from site'
    """
    tasks = []
    # #1 in-dms
    if not df_in.empty and "Photo Count" in df_in.columns:
        for _, row in df_in.iterrows():
            pc = row.get("Photo Count", "0")
            try:
                pc_val = float(pc)
            except:
                pc_val = 0
            if pc_val>1:
                # check sites
                for sitecol in ["is_on_cars", "is_on_autotrader", "is_on_pmgWeb"]:
                    if row.get(sitecol,"No")=="No":
                        tasks.append({
                            "Task": "Need to fix listing",
                            "Stock Number": row["Stock Number"],
                            "Notes": f"PhotoCount>1 but missing on {sitecol.replace('is_on_','')}"
                        })
                        break

    # #2 df_rm => remove from site
    if not df_rm.empty:
        for _, row in df_rm.iterrows():
            tasks.append({
                "Task":"Remove from site",
                "Stock Number": row["Stock Number"],
                "Notes":"Website-only stock"
            })

    df = pd.DataFrame(tasks, columns=["Task","Stock Number","Notes"])
    return df


def main():
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    print("[stockgpt] Reading data from src/ ...")

    # 1) read dms
    df_dms = read_dms_data()
    # 2) read website sets
    at_stk, at_prices = read_autotrader_data()
    cars_stk, cars_prices = read_cars_data()
    pmg_stk, pmg_prices = read_pmg_web_data()

    # 3) build master
    df_master = build_master_df(df_dms, at_stk, at_prices, cars_stk, cars_prices, pmg_stk, pmg_prices)

    # 4) separate => in_dms & recognized prefix => each prefix sheet
    #    also in_dms=False => 'to_be_removed'
    recognized_prefixes = tuple(DEALER_PREFIXES.values())

    # to_be_removed => website-only => in_dms=False + recognized prefix
    mask_removed = (df_master["in_dms"]==False) & \
                   (df_master["Stock Number"].str.startswith(recognized_prefixes, na=False))
    df_removed = df_master[mask_removed].copy()

    # keep only 6 columns
    rm_cols = [
        "Stock Number",
        "is_on_cars","cars_price",
        "is_on_autotrader","autotrader_price",
        "is_on_pmgWeb"
    ]
    df_removed = df_removed.reindex(columns=rm_cols)

    # 5) generate to_dos
    # but we also need in_dms recognized
    mask_in = (df_master["in_dms"]==True) & \
              (df_master["Stock Number"].str.startswith(recognized_prefixes, na=False))
    df_in_dms = df_master[mask_in].copy()

    df_todos = generate_todos(df_in_dms, df_removed)

    # 6) create workbook, each prefix => separate sheet
    out_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # For each recognized prefix => create a sheet with that subset
    for dealer_name, prefix_val in DEALER_PREFIXES.items():
        # subset
        subset = df_in_dms[df_in_dms["Stock Number"].str.startswith(prefix_val, na=False)].copy()
        if subset.empty:
            continue
        ws = wb.create_sheet(dealer_name)
        # write subset to xlsx
        # we can just do a quick df->ws approach or do code below:
        from openpyxl.utils.dataframe import dataframe_to_rows

        rows = dataframe_to_rows(subset, index=False, header=True)
        for r_idx, row_data in enumerate(rows, start=1):
            ws.append(row_data)

        # color code
        color_code_worksheet(ws, subset)
        # convert to table + auto-size
        auto_size_and_table(ws, subset, table_name=f"{dealer_name}Table")

    # "to_be_removed" => single sheet
    ws_removed = wb.create_sheet("to_be_removed")
    from openpyxl.utils.dataframe import dataframe_to_rows
    rows_rm = dataframe_to_rows(df_removed, index=False, header=True)
    for r_idx, row_data in enumerate(rows_rm, start=1):
        ws_removed.append(row_data)
    color_code_worksheet(ws_removed, df_removed)
    auto_size_and_table(ws_removed, df_removed, table_name="RemovedTable")

    # "to_dos" => single sheet
    ws_todos = wb.create_sheet("to_dos")
    rows_td = dataframe_to_rows(df_todos, index=False, header=True)
    for r_idx, row_data in enumerate(rows_td, start=1):
        ws_todos.append(row_data)
    # not doing color code for to_dos, but we do table + auto-size
    auto_size_and_table(ws_todos, df_todos, table_name="ToDosTable")

    wb.save(out_path)
    print(f"[stockgpt] Wrote {out_path} with sheets => one sheet per prefix, plus to_be_removed & to_dos.")


# Color-coded convenience
def color_code_worksheet(ws, df):
    """
    Color-code rows:
      - GREEN => all 3 site columns == "Yes"
      - RED => Photo Count>1 & any site col == "No"
    """
    rows = len(df)
    cols = len(df.columns)
    if rows<1 or cols<1:
        return

    # A2 => last col => last row
    from openpyxl.utils import get_column_letter
    rng = f"A2:{get_column_letter(cols)}{rows+1}"

    # GREEN => all sites == Yes
    try:
        c1 = df.columns.get_loc("is_on_cars")+1
        c2 = df.columns.get_loc("is_on_autotrader")+1
        c3 = df.columns.get_loc("is_on_pmgWeb")+1
    except KeyError:
        # no site columns => skip
        return
    else:
        col1 = get_column_letter(c1)
        col2 = get_column_letter(c2)
        col3 = get_column_letter(c3)
        fill_g = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        formula_g = f'AND(${col1}2="Yes", ${col2}2="Yes", ${col3}2="Yes")'
        rule_g = FormulaRule(formula=[formula_g], fill=fill_g)
        ws.conditional_formatting.add(rng, rule_g)

    # RED => PhotoCount>1 & any site == No
    # check if PhotoCount in df
    if "Photo Count" in df.columns:
        pc_ix = df.columns.get_loc("Photo Count")+1
        pcC = get_column_letter(pc_ix)
        # site cols again
        fill_r = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        formula_r = f'AND(${pcC}2>1, OR(${col1}2="No", ${col2}2="No", ${col3}2="No"))'
        rule_r = FormulaRule(formula=[formula_r], fill=fill_r)
        ws.conditional_formatting.add(rng, rule_r)

def auto_size_and_table(ws, df, table_name="DataTable"):
    """
    Turn the data region into a table & auto-size columns
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    row_count = len(df)
    col_count = len(df.columns)
    if row_count<1 or col_count<1:
        return

    last_col_letter = get_column_letter(col_count)
    ref = f"A1:{last_col_letter}{row_count+1}"  # +1 for header
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # approximate auto-size
    for i, col_name in enumerate(df.columns, start=1):
        sample_vals = df[col_name].astype(str).head(50).tolist()
        hdr_len = len(col_name)
        avg_len = sum(len(x) for x in sample_vals)/max(len(sample_vals),1)
        width = int(max(hdr_len, avg_len)) + 2
        ws.column_dimensions[get_column_letter(i)].width = width


if __name__=="__main__":
    main()
