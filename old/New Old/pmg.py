#!/usr/bin/env python3
"""
pmg.py

Styling & Tables Example:

1) Reads DMS (pinnacle columns) + website sets/dicts.
2) Builds a "master" DataFrame (DMS + website-only). 
   - 'in_dms' => True if from DMS
   - is_on_cars, cars_price, ...
   - if StockNumber starts with "UE" => is_on_cars="Yes"
3) Reorders columns so Photo Count, Internet Price, site columns go at the end.
4) Splits into:
   - Dealer sheets => in_dms=True & recognized prefix
   - "to_be_removed" => in_dms=False & recognized prefix (6 columns only)
5) Applies multiple conditional formats:
   a) Green fill if all site columns == "Yes" 
   b) Red fill if Photo Count > 1 & (any site column == "No")
6) Creates an Excel Table (so you can sort) on each sheet & auto-resizes columns.

Usage:
  python pmg.py
"""

import os
import warnings
import pandas as pd

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utilities import (
    read_csv_with_sep_check,
    clean_dataframe,
    col_index_to_excel_col_name
)

# Suppress openpyxl's "no default style" warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ------------------- CONFIG -------------------
PINNACLE_SCHEMA = [
    "Stock Number",
    "Make",
    "Model",
    "Specification",
    "Colour",
    "Registration Date",
    "VIN",
    "Odometer",
    "Photo Count",
    "Selling Price",
    "Stand In Value",
    "Internet Price",
    "Date In Stock",
    "Original Group Date In Stock",
    "Stock Days",
    "Branch",
    "Location",
    "Body Style",
    "Fuel Type",
    "Transmission",
    "Customer Ordered",
    "Profiles",
]

DEALER_PREFIXES = {
    "Ford_Nelspruit":   "UF",
    "Ford_Mazda":       "UG",
    "Produkta_Nissan":  "UA",
    "Suzuki_Nelspruit": "UE",  # forced is_on_cars="Yes"
    "Ford_Malalane":    "US",
}


# -------------- STEP A: READ DMS INTO DICT --------------
def extract_dms_dict(dms_file: str) -> dict:
    """Reads pmg_dms_data.csv -> {stockNum -> rowData} with only PINNACLE_SCHEMA."""
    if not os.path.isfile(dms_file):
        print(f"[WARN] Missing DMS file: {dms_file}")
        return {}

    # Peek columns
    temp_cols = pd.read_csv(dms_file, nrows=1).columns.tolist()
    wanted_cols = []
    for col in PINNACLE_SCHEMA:
        if col == "Customer Ordered" and "Customer Order" in temp_cols:
            wanted_cols.append("Customer Order")
        elif col in temp_cols:
            wanted_cols.append(col)

    df = pd.read_csv(dms_file, usecols=wanted_cols)
    if "Customer Order" in df.columns:
        df.rename(columns={"Customer Order": "Customer Ordered"}, inplace=True)

    # Reorder & strip
    final_cols = [c for c in PINNACLE_SCHEMA if c in df.columns]
    df = df[final_cols]
    for c in df.select_dtypes(include=['object']).columns:
        df[c] = df[c].astype(str).str.strip()

    # Convert to dict
    dms_map = {}
    if "Stock Number" not in df.columns:
        return dms_map

    for _, row in df.iterrows():
        sn = row["Stock Number"]
        if sn:
            dms_map[sn] = dict(row)
    return dms_map


# -------------- STEP B: READ WEBSITE DATA --------------
def read_autotrader_data(src_folder: str):
    stks, prices = set(), {}
    for sub in os.listdir(src_folder):
        path_sub = os.path.join(src_folder, sub)
        if not os.path.isdir(path_sub):
            continue
        csv_file = os.path.join(path_sub, "autotrader.csv")
        if not os.path.isfile(csv_file):
            continue

        df = read_csv_with_sep_check(csv_file)
        df.columns = df.columns.str.strip()
        if "StockNumber" in df.columns:
            df.rename(columns={"StockNumber": "Stock Number"}, inplace=True)

        price_col = next((c for c in ["PriceFormatted", "Price", "price"] if c in df.columns), None)

        if "Stock Number" in df.columns:
            for _, row in df.iterrows():
                sn = str(row["Stock Number"]).strip()
                if sn:
                    stks.add(sn)
                    p = str(row[price_col]).strip() if price_col else ""
                    prices[sn] = p
    return stks, prices


def read_cars_data(src_folder: str):
    stks, prices = set(), {}
    for sub in os.listdir(src_folder):
        path_sub = os.path.join(src_folder, sub)
        if not os.path.isdir(path_sub):
            continue
        xlsx_file = os.path.join(path_sub, "cars.xlsx")
        if not os.path.isfile(xlsx_file):
            continue

        df = pd.read_excel(xlsx_file)
        df.columns = df.columns.str.strip()
        if "Reference" in df.columns:
            df.rename(columns={"Reference": "Stock Number"}, inplace=True)

        price_col = next((c for c in ["Price", "price"] if c in df.columns), None)

        if "Stock Number" in df.columns:
            for _, row in df.iterrows():
                sn = str(row["Stock Number"]).strip()
                if sn:
                    stks.add(sn)
                    p = str(row[price_col]).strip() if price_col else ""
                    prices[sn] = p
    return stks, prices


def read_pmg_web_data(src_folder: str):
    csv_file = os.path.join(src_folder, "pmg_web_data.csv")
    if not os.path.isfile(csv_file):
        return set(), {}

    df = read_csv_with_sep_check(csv_file)
    df.columns = df.columns.str.strip()
    if "SKU" in df.columns:
        df.rename(columns={"SKU": "Stock Number"}, inplace=True)

    price_col = next((c for c in ["Regular price", "Regular Price", "price", "Price"] if c in df.columns), None)

    stks, prices = set(), {}
    if "Stock Number" in df.columns:
        for _, row in df.iterrows():
            sn = str(row["Stock Number"]).strip()
            if sn:
                stks.add(sn)
                p = str(row[price_col]).strip() if price_col else ""
                prices[sn] = p
    return stks, prices


# -------------- STEP C: BUILD MASTER DATAFRAME --------------
def build_master_dataframe(dms_map: dict,
                           at_set: set, at_prices: dict,
                           cars_set: set, cars_prices: dict,
                           pmg_set: set, pmg_prices: dict) -> pd.DataFrame:
    """
    Union of DMS + website stock. Each row => 'in_dms', site columns, etc.
    If prefix=UE => is_on_cars="Yes".
    """
    all_stocks = set(dms_map.keys()) | at_set | cars_set | pmg_set
    rows = []

    for sn in sorted(all_stocks):
        row = {}
        # DMS columns if present
        if sn in dms_map:
            row.update(dms_map[sn])
            row["in_dms"] = True
        else:
            for c in PINNACLE_SCHEMA:
                row[c] = ""
            row["in_dms"] = False

        # Cars
        if sn in cars_set:
            row["is_on_cars"] = "Yes"
            row["cars_price"] = cars_prices.get(sn, "")
        else:
            row["is_on_cars"] = "No"
            row["cars_price"] = ""

        # Autotrader
        if sn in at_set:
            row["is_on_autotrader"] = "Yes"
            row["autotrader_price"] = at_prices.get(sn, "")
        else:
            row["is_on_autotrader"] = "No"
            row["autotrader_price"] = ""

        # pmg web
        if sn in pmg_set:
            row["is_on_pmgWeb"] = "Yes"
            row["pmg_web_price"] = pmg_prices.get(sn, "")
        else:
            row["is_on_pmgWeb"] = "No"
            row["pmg_web_price"] = ""

        # Suzuki => prefix "UE" => is_on_cars="Yes"
        if sn.startswith("UE"):
            row["is_on_cars"] = "Yes"

        row["Stock Number"] = sn
        rows.append(row)

    return pd.DataFrame(rows)


def reorder_final_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Put main Pinnacle columns up front (plus 'in_dms'),
    then Photo Count, Internet Price, site columns at the end.
    """
    existing = df.columns.tolist()
    front_cols = [
        "Stock Number", "Make", "Model", "Specification", "Colour",
        "Registration Date", "VIN", "Odometer", "Selling Price",
        "Stand In Value", "Date In Stock", "Original Group Date In Stock",
        "Stock Days", "Branch", "Location", "Body Style", "Fuel Type",
        "Transmission", "Customer Ordered", "Profiles", "in_dms"
    ]
    end_cols = [
        "Photo Count", "Internet Price",
        "is_on_cars", "cars_price",
        "is_on_autotrader", "autotrader_price",
        "is_on_pmgWeb", "pmg_web_price"
    ]

    final_order = [c for c in front_cols if c in existing] + [c for c in end_cols if c in existing]
    df = df.reindex(columns=final_order)
    return clean_dataframe(df)


# -------------- STEP D: STYLING HELPERS --------------
def auto_size_columns(sheet, df: pd.DataFrame):
    """
    Approximate "auto size" for each column:
    - width = max(len(header), average len of sample rows) + a bit
    """
    from openpyxl.utils import get_column_letter

    for col_idx, col_name in enumerate(df.columns, start=1):
        # We'll check the length of the column name
        header_len = len(col_name)

        # sample a few rows in data for length
        sample_values = df[col_name].astype(str).head(50).tolist()  # sample up to 50
        avg_len = sum(len(v) for v in sample_values) / max(len(sample_values), 1)

        # pick bigger of header_len or avg_len
        best_len = int(max(header_len, avg_len)) + 2  # add padding
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = best_len


def create_excel_table(sheet, df: pd.DataFrame, table_name="DataTable"):
    """
    Insert an Excel Table so data is easily sortable/filterable.
    We'll use a medium style. 
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Table reference => "A1" through last column & row
    rows = df.shape[0]
    cols = df.shape[1]
    if rows < 1 or cols < 1:
        return  # no table if no data

    last_col_letter = col_index_to_excel_col_name(cols - 1)
    ref = f"A1:{last_col_letter}{rows + 1}"  # +1 for header row

    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)


def apply_conditional_formatting(sheet, df: pd.DataFrame):
    """
    1) Green fill if is_on_cars/is_on_autotrader/is_on_pmgWeb all "Yes"
    2) Red fill if Photo Count>1 & any site col== "No"
    """
    all_cols = df.columns.tolist()

    # 1) Green fill => all three site columns = "Yes"
    try:
        cars_ix = all_cols.index("is_on_cars")
        auto_ix = all_cols.index("is_on_autotrader")
        pmg_ix  = all_cols.index("is_on_pmgWeb")
    except ValueError:
        pass
    else:
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        c_cars = col_index_to_excel_col_name(cars_ix)
        c_auto = col_index_to_excel_col_name(auto_ix)
        c_pmg  = col_index_to_excel_col_name(pmg_ix)
        max_row = len(df) + 1
        max_col_letter = col_index_to_excel_col_name(len(all_cols) - 1)
        rng = f"A2:{max_col_letter}{max_row}"
        formula = f'AND(${c_cars}2="Yes", ${c_auto}2="Yes", ${c_pmg}2="Yes")'
        rule_green = FormulaRule(formula=[formula], fill=green_fill, stopIfTrue=False)
        sheet.conditional_formatting.add(rng, rule_green)

    # 2) Red fill => if Photo Count>1 & (any site col == "No")
    #    => formula example: AND($PhotoCountCol2>1, OR($c_cars2="No", $c_auto2="No", $c_pmg2="No"))
    try:
        photo_ix = all_cols.index("Photo Count")
    except ValueError:
        pass
    else:
        # We only do this if the site columns exist, or else skip
        if "is_on_cars" in all_cols and "is_on_autotrader" in all_cols and "is_on_pmgWeb" in all_cols:
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            c_photo = col_index_to_excel_col_name(photo_ix)
            c_cars = col_index_to_excel_col_name(all_cols.index("is_on_cars"))
            c_auto = col_index_to_excel_col_name(all_cols.index("is_on_autotrader"))
            c_pmg  = col_index_to_excel_col_name(all_cols.index("is_on_pmgWeb"))

            max_row = len(df) + 1
            max_col_letter = col_index_to_excel_col_name(len(all_cols) - 1)
            rng = f"A2:{max_col_letter}{max_row}"

            # Photo>1 => ($Photo>=2 basically)
            # any site col == "No" => OR($c_cars2="No", $c_auto2="No", $c_pmg2="No")
            formula_red = (
                f'AND(${c_photo}2>1, OR(${c_cars}2="No", ${c_auto}2="No", ${c_pmg}2="No"))'
            )

            rule_red = FormulaRule(formula=[formula_red], fill=red_fill, stopIfTrue=False)
            sheet.conditional_formatting.add(rng, rule_red)


def style_sheet(writer, sheet_name, df: pd.DataFrame):
    """
    1) Create an Excel table so data is easily sortable
    2) Auto-size columns
    3) Apply conditional formatting
    """
    sheet = writer.sheets[sheet_name]

    # Insert table
    create_excel_table(sheet, df, table_name=f"{sheet_name}Table")

    # Approx auto-size
    auto_size_columns(sheet, df)

    # Conditionals (green & red)
    apply_conditional_formatting(sheet, df)


# -------------- STEP E: MAIN --------------
def main():
    src_folder = "src"
    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    # 1) DMS -> dict
    dms_file = os.path.join(src_folder, "pmg_dms_data.csv")
    dms_map = extract_dms_dict(dms_file)

    # 2) Website sets/dicts
    at_set, at_prices   = read_autotrader_data(src_folder)
    cars_set, cars_prices = read_cars_data(src_folder)
    pmg_set, pmg_prices   = read_pmg_web_data(src_folder)

    # 3) Master
    df_master = build_master_dataframe(
        dms_map, at_set, at_prices, cars_set, cars_prices, pmg_set, pmg_prices
    )
    df_master = reorder_final_columns(df_master)

    # 4) Split => in_dms vs. to_be_removed
    recognized_prefixes = tuple(DEALER_PREFIXES.values())  # e.g. ("UF","UG","UA","UE","US")

    df_to_remove = df_master[
        (~df_master["in_dms"]) &
        df_master["Stock Number"].str.startswith(recognized_prefixes, na=False)
    ].copy()

    df_in_dms = df_master[
        (df_master["in_dms"]) &
        df_master["Stock Number"].str.startswith(recognized_prefixes, na=False)
    ].copy()

    # "to_be_removed" => 6 columns only
    remove_cols = [
        "Stock Number",
        "is_on_cars", "cars_price",
        "is_on_autotrader", "autotrader_price",
        "is_on_pmgWeb"
    ]

    # 5) Write sheets
    out_file = os.path.join(output_folder, "dealers_by_prefix.xlsx")
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        sheet_count = 0

        # A) Dealer sheets first
        for dealer_name, prefix in DEALER_PREFIXES.items():
            df_sub = df_in_dms[df_in_dms["Stock Number"].str.startswith(prefix)].copy()
            if df_sub.empty:
                continue
            df_sub.to_excel(writer, sheet_name=dealer_name, index=False)
            sheet_count += 1

        # B) Then "to_be_removed" last
        if not df_to_remove.empty:
            df_rem = df_to_remove.reindex(columns=remove_cols)
            df_rem = clean_dataframe(df_rem)  # remove fully blank lines/cols
            df_rem.to_excel(writer, sheet_name="to_be_removed", index=False)
            sheet_count += 1

        # Now style each sheet with table, conditional formats, etc.
        for sheet_name in writer.sheets.keys():
            # We load the DataFrame from the workbook by name
            # but we already have df_sub or df_rem from memory.
            # We'll simply reconstruct them:
            # Actually let's do it simpler: read back from writer.sheets is not trivial,
            # so let's map them ourselves:
            if sheet_name == "to_be_removed":
                style_sheet(writer, sheet_name, df_rem)
            else:
                # find which prefix?
                prefix = DEALER_PREFIXES.get(sheet_name, None)
                if prefix:
                    # the subset
                    df_sub = df_in_dms[df_in_dms["Stock Number"].str.startswith(prefix)].copy()
                    style_sheet(writer, sheet_name, df_sub)

    print(f"[INFO] Wrote {sheet_count} sheets to {out_file} with styling & table.")
    print("Dealer sheets first, then 'to_be_removed' last.")


if __name__ == "__main__":
    main()
