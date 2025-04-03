# excel_report.py
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
import warnings

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

def apply_table(ws, table_name):
    if ws.max_row > 1 and ws.max_column > 1:
        ref = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

def normalize_stock_column(df, col_candidates):
    for col in col_candidates:
        if col in df.columns:
            df = df.rename(columns={col: "Stock Number"})
            df["Stock Number"] = df["Stock Number"].astype(str).str.strip().str.upper()
            break
    return df

def write_master_excel(dms_data: pd.DataFrame,
                       pmg_web_data: pd.DataFrame,
                       dealership_data: dict,
                       output_path: Path):
    """Create a master Excel file with all source data in ordered sheets."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    dms_columns = [
        "Stock Number", "Make", "Model", "Specification", "Odometer", "Registration Date",
        "VIN", "Customer Order", "Photo Count", "Selling Price", "Stand In Value",
        "Stock Days", "Internet Price", "Vehicle Code"
    ]

    dealership_prefixes = [
        ("FordNelspruit", "UF"),
        ("MazdaNelspruit", "UG"),
        ("ProduktaNissan", "UA"),
        ("SuzukiNelspruit", "UE"),
        ("FordMalalane", "US"),
    ]

    dms_data = normalize_stock_column(dms_data, ["Stock Number"])
    pmg_web_data = normalize_stock_column(pmg_web_data, ["SKU"])

    autotrader_combined = pd.concat(
        [normalize_stock_column(data['autotrader'], ["Stock Number", "Reference", "StockNumber", "Ref"])
         for data in dealership_data.values() if 'autotrader' in data],
        ignore_index=True
    )
    cars_combined = pd.concat(
        [normalize_stock_column(data['cars'], ["Stock Number", "Reference", "StockNumber", "Ref"])
         for data in dealership_data.values() if 'cars' in data],
        ignore_index=True
    )

    for dealership, prefix in dealership_prefixes:
        df_filtered = dms_data[dms_data['Stock Number'].astype(str).str.startswith(prefix)].copy()
        df_filtered = df_filtered[[col for col in dms_columns if col in df_filtered.columns]]

        df_filtered["on_cars"] = df_filtered["Stock Number"].isin(cars_combined.get("Stock Number", pd.Series())).astype(bool)
        df_filtered["on_autotrader"] = df_filtered["Stock Number"].isin(autotrader_combined.get("Stock Number", pd.Series())).astype(bool)
        df_filtered["on_pmgweb"] = df_filtered["Stock Number"].isin(pmg_web_data.get("Stock Number", pd.Series())).astype(bool)

        sheet_name = f"{dealership} DMS"
        ws = wb.create_sheet(sheet_name[:31])
        for row in dataframe_to_rows(df_filtered, index=False, header=True):
            ws.append(row)
        apply_table(ws, sheet_name.replace(" ", "_"))
        autofit_columns(ws)

    ws_auto = wb.create_sheet("AutoTrader")
    for row in dataframe_to_rows(autotrader_combined, index=False, header=True):
        ws_auto.append(row)
    apply_table(ws_auto, "AutoTrader")
    autofit_columns(ws_auto)

    ws_cars = wb.create_sheet("Cars")
    for row in dataframe_to_rows(cars_combined, index=False, header=True):
        ws_cars.append(row)
    apply_table(ws_cars, "Cars")
    autofit_columns(ws_cars)

    ws_web = wb.create_sheet("PMG_Web")
    for row in dataframe_to_rows(pmg_web_data, index=False, header=True):
        ws_web.append(row)
    apply_table(ws_web, "PMG_Web")
    autofit_columns(ws_web)

    # Add separator sheet
    wb.create_sheet("-->")

    # Vehicles to be removed: listed online but not in DMS
    all_dms_stock = set(dms_data["Stock Number"].unique())
    remove_autotrader = autotrader_combined[~autotrader_combined["Stock Number"].isin(all_dms_stock)].copy()
    remove_cars = cars_combined[~cars_combined["Stock Number"].isin(all_dms_stock)].copy()
    remove_web = pmg_web_data[~pmg_web_data["Stock Number"].isin(all_dms_stock)].copy()

    if not remove_autotrader.empty:
        ws_rem_auto = wb.create_sheet("To_Remove_AutoTrader")
        for row in dataframe_to_rows(remove_autotrader, index=False, header=True):
            ws_rem_auto.append(row)
        apply_table(ws_rem_auto, "To_Remove_AutoTrader")
        autofit_columns(ws_rem_auto)

    if not remove_cars.empty:
        ws_rem_cars = wb.create_sheet("To_Remove_Cars")
        for row in dataframe_to_rows(remove_cars, index=False, header=True):
            ws_rem_cars.append(row)
        apply_table(ws_rem_cars, "To_Remove_Cars")
        autofit_columns(ws_rem_cars)

    if not remove_web.empty:
        ws_rem_web = wb.create_sheet("To_Remove_PMGWeb")
        for row in dataframe_to_rows(remove_web, index=False, header=True):
            ws_rem_web.append(row)
        apply_table(ws_rem_web, "To_Remove_PMGWeb")
        autofit_columns(ws_rem_web)

    # Vehicles to upload to PMG Web: listed on AutoTrader or Cars but not on PMG Web
    to_upload_web = dms_data[
        (
            dms_data["Stock Number"].isin(autotrader_combined.get("Stock Number", pd.Series())) |
            dms_data["Stock Number"].isin(cars_combined.get("Stock Number", pd.Series()))
        ) & ~dms_data["Stock Number"].isin(pmg_web_data.get("Stock Number", pd.Series()))
    ].copy()

    if not to_upload_web.empty:
        ws_up_web = wb.create_sheet("Upload_to_PMGWeb")
        for row in dataframe_to_rows(to_upload_web, index=False, header=True):
            ws_up_web.append(row)
        apply_table(ws_up_web, "Upload_to_PMGWeb")
        autofit_columns(ws_up_web)

    wb.save(output_path)
