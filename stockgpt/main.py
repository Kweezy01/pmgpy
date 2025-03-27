#!/usr/bin/env python3
"""
main.py - Generates stockgpt.xlsx with:
 - Dealer-specific sheets (based on stock number prefix)
 - to_be_removed: Vehicles listed on web but not in DMS
 - to_dos: Vehicles missing from one or more websites
 - corporate_report: Summary stats & graphs

Ready for:
 - MySQL integration (future)
 - UI tracking (future)
"""

import os
import pandas as pd
import warnings
from openpyxl.chart import BarChart, PieChart, Reference
from utilities import clean_dataframe
from openpyxl.styles import PatternFill
from .data_readers import read_dms_dict, read_autotrader_data, read_cars_data, read_pmg_web_data
from .transformations import build_master_df, reorder_final_columns, generate_todos
from .formatting import style_sheet, auto_size_columns

# Suppress OpenPyXL formatting warnings
warnings.simplefilter("ignore", UserWarning)

# Dealer prefixes map
DEALER_PREFIXES = {
    "Ford_Nelspruit":   "UF",
    "Ford_Mazda":       "UG",
    "Produkta_Nissan":  "UA",
    "Suzuki_Nelspruit": "UE",
    "Ford_Malalane":    "US",
}

def generate_corporate_report(writer, df_master):
    ws = writer.book.create_sheet("corporate_report")

    # Key stats
    counts = {
        "DMS Stock": df_master[df_master["in_dms"]].shape[0],
        "Cars.co.za": df_master[df_master["is_on_cars"] == "Yes"].shape[0],
        "AutoTrader": df_master[df_master["is_on_autotrader"] == "Yes"].shape[0],
        "PMG Web": df_master[df_master["is_on_pmgWeb"] == "Yes"].shape[0],
        "To Be Removed": df_master[~df_master["in_dms"]].shape[0],
    }

    ws.append(["Corporate Vehicle Report"])
    ws.append([""])
    for k, v in counts.items():
        ws.append([k, v])
    ws.append([""])
    ws.append(["Dealership", "Stock Count"])

    row_start = ws.max_row + 1
    for dealer, prefix in DEALER_PREFIXES.items():
        count = df_master[df_master["Stock Number"].str.startswith(prefix)].shape[0]
        ws.append([dealer, count])
    row_end = ws.max_row

    # Bar chart
    bar = BarChart()
    bar.title = "Vehicles Per Dealership"
    bar_data = Reference(ws, min_col=2, min_row=row_start, max_row=row_end)
    bar_labels = Reference(ws, min_col=1, min_row=row_start + 1, max_row=row_end)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_labels)
    bar.width, bar.height = 12, 6
    ws.add_chart(bar, "E5")

    # Pie chart
    pie = PieChart()
    pie.title = "Web Presence Distribution"
    pie_data = Reference(ws, min_col=2, min_row=3, max_row=5)
    pie_labels = Reference(ws, min_col=1, min_row=3, max_row=5)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    ws.add_chart(pie, "E15")

    auto_size_columns(ws, df_master)

def main():
    src = "src"
    out = "output"
    os.makedirs(out, exist_ok=True)

    print("[INFO] Reading DMS...")
    dms = read_dms_dict(os.path.join(src, "pmg_dms_data.csv"))
    print(f"[INFO] DMS vehicles: {len(dms)}")

    print("[INFO] Reading website data...")
    at_set, at_prices = read_autotrader_data(src)
    cars_set, cars_prices = read_cars_data(src)
    pmg_set, pmg_prices = read_pmg_web_data(src)

    print("[INFO] Building dataset...")
    df_master = reorder_final_columns(build_master_df(
        dms, at_set, at_prices, cars_set, cars_prices, pmg_set, pmg_prices
    ))

    with pd.ExcelWriter(os.path.join(out, "stockgpt.xlsx"), engine="openpyxl") as writer:
        # Dealer sheets
        for dealer, prefix in DEALER_PREFIXES.items():
            df = df_master[df_master["Stock Number"].str.startswith(prefix)]
            if not df.empty:
                df.to_excel(writer, sheet_name=dealer, index=False)
                style_sheet(writer, dealer, df)

        # to_be_removed sheet
        df_rem = df_master[~df_master["in_dms"]][
            ["Stock Number", "is_on_cars", "cars_price", "is_on_autotrader", "autotrader_price", "is_on_pmgWeb"]
        ].copy()
        df_rem["Done?"] = ""
        df_rem.to_excel(writer, sheet_name="to_be_removed", index=False)
        style_sheet(writer, "to_be_removed", df_rem)

        # to_dos sheet
        df_todo = generate_todos(df_master, df_rem)
        df_todo["Done?"] = ""
        df_todo.to_excel(writer, sheet_name="to_dos", index=False)
        style_sheet(writer, "to_dos", df_todo)

        # Corporate report
        generate_corporate_report(writer, df_master)

    print("[DONE] Excel generated with all sheets & corporate report.")

if __name__ == "__main__":
    main()
