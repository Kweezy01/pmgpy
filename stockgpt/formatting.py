from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference, PieChart

def auto_size_columns(sheet, df):
    for i, col in enumerate(df.columns, 1):
        max_len = max(df[col].astype(str).str.len().mean(), len(str(col))) + 2
        col_letter = get_column_letter(i)
        sheet.column_dimensions[col_letter].width = max_len

def create_excel_table(sheet, df, name="Table1"):
    if df.empty:
        return
    end_col = get_column_letter(len(df.columns))
    ref = f"A1:{end_col}{len(df)+1}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)

def apply_conditional_formatting(sheet, df):
    if df.empty:
        return

    if not all(c in df.columns for c in ["is_on_cars", "is_on_autotrader", "is_on_pmgWeb"]):
        return

    max_row = len(df) + 1
    col_cars = get_column_letter(df.columns.get_loc("is_on_cars") + 1)
    col_auto = get_column_letter(df.columns.get_loc("is_on_autotrader") + 1)
    col_pmg = get_column_letter(df.columns.get_loc("is_on_pmgWeb") + 1)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    formula_green = f'=AND(${col_cars}2="Yes", ${col_auto}2="Yes", ${col_pmg}2="Yes")'
    formula_red = f'=AND(COUNTIF(${col_cars}2:${col_pmg}2, "Yes")>=1, COUNTIF(${col_cars}2:${col_pmg}2, "No")>=1)'

    sheet.conditional_formatting.add(f"A2:{col_pmg}{max_row}", FormulaRule(formula=[formula_green], fill=green_fill))
    sheet.conditional_formatting.add(f"A2:{col_pmg}{max_row}", FormulaRule(formula=[formula_red], fill=red_fill))

    if "Done?" in df.columns:
        col_done = get_column_letter(df.columns.get_loc("Done?") + 1)
        done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        formula_done = f'=${col_done}2=TRUE'
        sheet.conditional_formatting.add(f"A2:{col_done}{max_row}", FormulaRule(formula=[formula_done], fill=done_fill))


def style_sheet(writer, sheet_name, df):
    sheet = writer.sheets[sheet_name]
    create_excel_table(sheet, df, name=f"{sheet_name}_table")
    auto_size_columns(sheet, df)
    apply_conditional_formatting(sheet, df)

def generate_corporate_report(writer, df_master):
    ws = writer.book.create_sheet("corporate_report")
    ws.append(["Corporate Vehicle Report"])
    ws.append([])
    fields = [
        ("DMS Stock", df_master[df_master["in_dms"]].shape[0]),
        ("Cars.co.za", df_master[df_master["is_on_cars"] == "Yes"].shape[0]),
        ("AutoTrader", df_master[df_master["is_on_autotrader"] == "Yes"].shape[0]),
        ("PMG Web", df_master[df_master["is_on_pmgWeb"] == "Yes"].shape[0]),
        ("To Be Removed", df_master[~df_master["in_dms"]].shape[0]),
    ]
    for f in fields:
        ws.append(list(f))
    ws.append([])
    ws.append(["Dealership", "Count"])

    prefixes = {
        "Ford_Nelspruit": "UF", "Ford_Mazda": "UG", "Produkta_Nissan": "UA",
        "Suzuki_Nelspruit": "UE", "Ford_Malalane": "US"
    }

    start = ws.max_row + 1
    for name, prefix in prefixes.items():
        count = df_master[df_master["Stock Number"].str.startswith(prefix)].shape[0]
        ws.append([name, count])
    end = ws.max_row

    bar = BarChart()
    data = Reference(ws, min_col=2, min_row=start, max_row=end)
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=end)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.title = "Stock by Dealership"
    bar.width, bar.height = 10, 6
    ws.add_chart(bar, "E5")

    pie = PieChart()
    pie_data = Reference(ws, min_col=2, min_row=3, max_row=6)
    pie_labels = Reference(ws, min_col=1, min_row=3, max_row=6)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    pie.title = "Online Stock Presence"
    ws.add_chart(pie, "E15")

    auto_size_columns(ws, df_master)
